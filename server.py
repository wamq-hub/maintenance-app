# -*- coding: utf-8 -*-
import sys
import os
import shutil
from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import base64, re
from datetime import datetime
import tempfile

app = Flask(__name__)
CORS(app)

TEMPLATE_PATH = "template.xlsx"
DEFAULT_LOGO = "assets/logo.png"

try:
    from PIL import Image as PILImage
except ImportError:
    PILImage = None

def check_template_permissions():
    """فحص وإصلاح صلاحيات ملف القالب"""
    if not os.path.exists(TEMPLATE_PATH):
        print(f"تحذير: ملف القالب غير موجود: {TEMPLATE_PATH}")
        return False
    
    # فحص صلاحيات القراءة
    if not os.access(TEMPLATE_PATH, os.R_OK):
        print(f"خطأ: لا توجد صلاحية قراءة لملف: {TEMPLATE_PATH}")
        return False
    
    # فحص صلاحيات الكتابة للمجلد
    template_dir = os.path.dirname(os.path.abspath(TEMPLATE_PATH))
    if not os.access(template_dir, os.W_OK):
        print(f"تحذير: لا توجد صلاحية كتابة في المجلد: {template_dir}")
    
    print(f"ملف القالب متاح: {TEMPLATE_PATH}")
    return True

def safe_load_template():
    """تحميل القالب بطريقة آمنة"""
    try:
        # تحميل القالب مباشرة في الذاكرة
        with open(TEMPLATE_PATH, 'rb') as f:
            template_data = f.read()
        wb = load_workbook(BytesIO(template_data))
        return wb
    except Exception as e:
        print(f"خطأ في تحميل القالب: {e}")
        raise
def safe_set(ws, cell_address, value):
    """تعيين قيمة للخلية مع دعم الخلايا المدمجة"""
    try:
        cell = ws[cell_address]
        if isinstance(cell, MergedCell):
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    top_left = merged_range.min_row, merged_range.min_col
                    ws.cell(row=top_left[0], column=top_left[1], value=value)
                    return
        else:
            cell.value = value
    except Exception as e:
        print(f"خطأ في تعيين القيمة للخلية {cell_address}: {e}")

def add_logo(ws, logo_path=None, data_url=None, anchor_cell="E5", width=120, height=120):
    """إدراج الشعار في الورقة"""
    img = None

    # 1) تفضيل data_url من العميل
    if data_url and isinstance(data_url, str) and data_url.startswith("data:image"):
        m = re.match(r"^data:image/[^;]+;base64,(.+)$", data_url)
        if m:
            try:
                raw = base64.b64decode(m.group(1))
                bio = BytesIO(raw)
                bio.seek(0)
                if PILImage:
                    try:
                        pil = PILImage.open(bio).convert("RGBA")
                        out = BytesIO()
                        pil.save(out, format="PNG")
                        out.seek(0)
                        img = XLImage(out)
                    except Exception as e:
                        print("فشل تحويل data_url إلى PNG:", e)
                        bio.seek(0)
                        try:
                            img = XLImage(bio)
                        except Exception as e2:
                            print("فشل XLImage من data_url:", e2)
                else:
                    try:
                        img = XLImage(bio)
                    except Exception as e:
                        print("Pillow مطلوب لبعض تنسيقات الصور:", e)
            except Exception as e:
                print("خطأ في معالجة data_url:", e)

    # 2) العودة للملف المحلي
    if img is None and logo_path and os.path.exists(logo_path):
        try:
            if PILImage:
                pil = PILImage.open(logo_path).convert("RGBA")
                out = BytesIO()
                pil.save(out, format="PNG")
                out.seek(0)
                img = XLImage(out)
            else:
                img = XLImage(logo_path)
        except Exception as e:
            print("فشل تحميل الشعار من الملف:", e)

    if img is None:
        print("لم يتم توفير صورة شعار صالحة")
        return

    try:
        img.width, img.height = width, height
        ws.add_image(img, anchor_cell)
        print(f"تم إدراج الشعار في {anchor_cell} بحجم: {width} x {height}")
    except Exception as e:
        print(f"خطأ في إدراج الشعار: {e}")

# خريطة الخلايا المحدثة حسب القالب المرفوع

CELL_MAP = {
    "request_id": "C8",           # رقم الطلب
    "maintenance_type": "F8",     # نوع الطلب
    "location": "I8",             # المبنى/الدور
    "priority": "D9",             # نوع الصيانة
    "requester_name": "D11",      # اسم صاحب الطلب
    "request_time": "D13",        # الوقت
    "request_date": "H13",        # التاريخ
    "fault_type": "C15",          # العطل
    "fault_desc": "A19",          # وصف العطل
    "technician_name": "C25",     # اسم الفني
    "execution_date": "H26",      # التاريخ (الفني)
    "supervisor_name": "C29",     # اسم المشرف
    "status": "C30",              # حالة الطلب
    "status_date": "H31",         # التاريخ (المشرف)
    "requester_name_2": "C34",    # اسم صاحب الطلب (مكرر)
    "is_fixed": "C35",            # هل تم إصلاح العطل
    "fixed_date": "H36",          # التاريخ (الإصلاح)
}

@app.route("/api/export-pdf", methods=["POST"])
def export_pdf():
    try:
        # نفس كود export_excel لكن نحول لـ PDF
        payload = request.get_json(force=True) or {}
        if "request_id" not in payload:
            return jsonify({"ok": False, "error": "request_id مطلوب"}), 400

        # تحميل وتعبئة القالب (نفس الكود السابق)
        with open(TEMPLATE_PATH, 'rb') as f:
            template_data = f.read()
        wb = load_workbook(BytesIO(template_data))
        ws = wb.active

        # تعبئة البيانات
        for key, cell_addr in CELL_MAP.items():
            value = payload.get(key, "")
            if value:
                safe_set(ws, cell_addr, value)

        # حفظ كـ Excel مؤقت
        temp_excel = BytesIO()
        wb.save(temp_excel)
        temp_excel.seek(0)

        # تحويل لـ PDF باستخدام LibreOffice
        import subprocess
        import tempfile
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            temp_file.write(temp_excel.read())
            temp_file.flush()
            
            # تحويل لـ PDF
            pdf_path = temp_file.name.replace('.xlsx', '.pdf')
            return jsonify({"ok": False, "error": "LibreOffice غير مثبت"}), 500
            
            # قراءة PDF
            with open(pdf_path, 'rb') as pdf_file:
                pdf_data = pdf_file.read()
            
            # حذف الملفات المؤقتة
            os.unlink(temp_file.name)
            os.unlink(pdf_path)

        filename = f"تقرير_صيانة_{payload.get('request_id')}.pdf"
        
        return send_file(
            BytesIO(pdf_data),
            as_attachment=True,
            download_name=filename,
            mimetype="application/pdf"
        )
        
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500
        
@app.route("/api/export-excel", methods=["POST"])
def export_excel():
    try:
        payload = request.get_json(force=True) or {}
        if "request_id" not in payload:
            return jsonify({"ok": False, "error": "request_id مطلوب"}), 400

        # فحص وجود القالب
        if not check_template_permissions():
            return jsonify({"ok": False, "error": "ملف القالب غير متاح أو لا توجد صلاحيات كافية"}), 500

        # تحميل القالب بطريقة آمنة
        try:
            wb = safe_load_template()
            ws = wb.active
        except Exception as e:
            return jsonify({"ok": False, "error": f"فشل في تحميل القالب: {str(e)}"}), 500

        print(f"تم تحميل القالب بنجاح. الأوراق المتاحة: {wb.sheetnames}")

        # إضافة الشعار
        logo_len = len(payload.get("logo_data_url") or "")
        print(f"طول logo_data_url: {logo_len} | وجود الشعار الافتراضي: {os.path.exists(DEFAULT_LOGO)}")

        add_logo(
            ws,
            logo_path=DEFAULT_LOGO if os.path.exists(DEFAULT_LOGO) else None,
            data_url=payload.get("logo_data_url"),
            anchor_cell="E1",  # تعديل الموقع حسب القالب
            width=100, height=80
        )

        # تعبئة البيانات
        for key, cell_addr in CELL_MAP.items():
            value = payload.get(key, "")
            if value:  # تعبئة فقط إذا كانت القيمة موجودة
                safe_set(ws, cell_addr, value)
                print(f"تم تعيين {key} = {value} في الخلية {cell_addr}")

        # حفظ الملف في الذاكرة
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # إنشاء اسم ملف آمن
        rid = str(payload.get("request_id"))
        safe_request_id = re.sub(r"[^A-Za-z0-9_\u0600-\u06FF-]+", "_", rid)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"تقرير_صيانة_{safe_request_id}_{timestamp}.xlsx"

        print(f"تم إنشاء التقرير بنجاح: {filename}")

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        
    except Exception as e:
        error_msg = f"خطأ في إنشاء التقرير: {str(e)}"
        print(error_msg)
        return jsonify({"ok": False, "error": error_msg}), 500

@app.route("/api/health", methods=["GET"])
def health_check():
    """فحص حالة الخادم"""
    template_status = check_template_permissions()
    return jsonify({
        "status": "ok",
        "template_available": template_status,
        "template_path": TEMPLATE_PATH,
        "logo_available": os.path.exists(DEFAULT_LOGO)
    })

if __name__ == "__main__":
    print("🚀 بدء تشغيل خادم تقارير الصيانة المحسن...")
    print("📋 فحص المتطلبات...")
    
    # فحص المتطلبات
    required_packages = ['openpyxl', 'flask', 'flask-cors']
    missing_packages = []
    
    for package in required_packages:
        try:
            __import__(package.replace('-', '_'))
        except ImportError:
            missing_packages.append(package)
    
    if missing_packages:
        print(f"❌ مكتبات مفقودة: {', '.join(missing_packages)}")
        print(f"👉 لتثبيت المكتبات: pip install {' '.join(missing_packages)}")
    else:
        print("✅ جميع المكتبات متوفرة")
    
    # فحص القالب
    if check_template_permissions():
        print("✅ ملف القالب متاح")
    else:
        print("❌ مشكلة في ملف القالب")
        print("💡 تأكد من وجود ملف template.xlsx في نفس مجلد الخادم")
    
    print("🌐 الخادم يعمل على: http://localhost:5000")
    print("🔍 للفحص: http://localhost:5000/api/health")
    
    app.run("0.0.0.0", 5000, debug=True)